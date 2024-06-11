import openpyxl as xl

wb = xl.load_workbook('printers.xlsx')
sheet = wb['INTERNAL']

for row in range(2, sheet.max_row +1):
    cell = sheet.cell(row, 3)
    if cell.value == 'Admin':
        newcell = sheet.cell(row, 14)
        newcell2 = sheet.cell(row, 15)
        newcell3 = sheet.cell(row, 16)
        newcell.value = 'HP LJ MFP E52645DN'
        newcell2.value = 'R29,599.99'
        newcell3.value = '879.99'
    elif cell.value == 'Embroidery':
        newcell = sheet.cell(row, 14)
        newcell2 = sheet.cell(row, 15)
        newcell3 = sheet.cell(row, 16)
        newcell.value = 'HP COLOUR LJ MFP E47528f'
        newcell2.value = 'R25,199.99'
        newcell3.value = '729.99'
    else:
        newcell = sheet.cell(row, 14)
        newcell2 = sheet.cell(row, 15)
        newcell3 = sheet.cell(row, 16)
        newcell.value = 'HP LJ MFP E42540f'
        newcell2.value = 'R19,599.99'
        newcell3.value = '529.99'
wb.save('printers2.xlsx')
